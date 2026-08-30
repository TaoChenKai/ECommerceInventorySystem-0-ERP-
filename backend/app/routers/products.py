import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from ..database import get_db
from ..models import Category, Spu, Sku, ProductImage, Unit, WeightUnit, AuditLog, User
from ..schemas import (CategoryCreate, SpuCreate, SpuUpdate, SpuOut,
                       UnitCreate, UnitUpdate, UnitOut)
from ..deps import get_current_user

router = APIRouter(prefix="/api", tags=["products"])


def add_log(db: Session, user: User, action: str, detail: str = ""):
    db.add(AuditLog(user_id=user.id, username=user.username, action=action, detail=detail))


def spu_to_out(spu: Spu) -> SpuOut:
    skus = sorted(spu.skus, key=lambda s: s.id)
    images = sorted(spu.images, key=lambda im: (0 if im.img_type == "main" else 1, im.sort, im.id))
    return SpuOut(
        id=spu.id,
        name=spu.name,
        code=spu.code or "",
        category_id=spu.category_id,
        category_name=spu.category.name if spu.category else None,
        unit=spu.unit,
        weight=spu.weight or 0,
        weight_unit=spu.weight_unit or "千克",
        designer=spu.designer or "",
        production_date=spu.production_date,
        material=spu.material or "",
        image_url=spu.image_url or "",
        remark=spu.remark or "",
        created_at=spu.created_at,
        skus=skus,
        sku_count=len(skus),
        total_stock=sum(s.stock or 0 for s in skus),
        images=images,
    )


def _sync_images(db: Session, spu: Spu, body_images: list):
    """全量同步商品图片：保留提交中带 id 的、删掉未提交的、新增无 id 的"""
    submitted = [im.id for im in body_images if im.id]
    for old in list(spu.images):
        if old.id not in submitted:
            db.delete(old)
    for idx, im in enumerate(body_images):
        sort = im.sort if im.sort is not None else idx
        if im.id:
            target = db.query(ProductImage).filter(
                ProductImage.id == im.id, ProductImage.spu_id == spu.id).first()
            if target:
                target.img_type = im.img_type or "main"
                target.url = im.url.strip()
                target.sort = sort
        elif im.url.strip():
            spu.images.append(ProductImage(
                img_type=im.img_type or "main",
                url=im.url.strip(),
                sort=sort,
            ))


def _ensure_unit_name(db: Session, model, name: str):
    """单位/重量单位名若不在字典中则自动补录（用户可随时自加，这里兜底）"""
    if not name:
        return
    exists = db.query(model).filter(model.name == name.strip()).first()
    if not exists:
        db.add(model(name=name.strip()))


def _resolve_category(db: Session, body) -> int | None:
    """分类解析：优先按 category_name 按名定位/新建；否则退回 category_id"""
    name = (getattr(body, "category_name", "") or "").strip()
    if name:
        cat = db.query(Category).filter(Category.name == name).first()
        if not cat:
            cat = Category(name=name)
            db.add(cat)
            db.flush()
        return cat.id
    return body.category_id or None


# ================= 分类 =================
@router.get("/categories")
def list_categories(db: Session = Depends(get_db), _u: User = Depends(get_current_user)):
    return db.query(Category).order_by(Category.id).all()


@router.post("/categories")
def create_category(body: CategoryCreate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    if db.query(Category).filter(Category.name == body.name.strip()).first():
        raise HTTPException(400, "分类已存在")
    c = Category(name=body.name.strip(), remark=body.remark.strip())
    db.add(c)
    db.commit()
    db.refresh(c)
    add_log(db, u, "新增分类", c.name)
    db.commit()
    db.refresh(c)
    return c


@router.delete("/categories/{cid}")
def delete_category(cid: int, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    c = db.query(Category).filter(Category.id == cid).first()
    if not c:
        raise HTTPException(404, "分类不存在")
    if db.query(Spu).filter(Spu.category_id == cid).first():
        raise HTTPException(400, "该分类下还有商品，不能删除")
    name = c.name
    db.delete(c)
    db.commit()
    add_log(db, u, "删除分类", name)
    db.commit()
    return {"ok": True}


# ================= 商品 SPU / SKU =================
@router.get("/spus")
def list_spus(keyword: str = "", category_id: int = 0,
              db: Session = Depends(get_db), _u: User = Depends(get_current_user)):
    q = db.query(Spu).options(joinedload(Spu.skus)).filter(Spu.deleted_at.is_(None)).order_by(Spu.id.desc())
    if keyword.strip():
        like = f"%{keyword.strip()}%"
        q = q.filter(or_(Spu.name.like(like), Spu.code.like(like)))
    if category_id:
        q = q.filter(Spu.category_id == category_id)
    return [spu_to_out(s) for s in q.all()]


@router.get("/spus/{spu_id}")
def get_spu(spu_id: int, db: Session = Depends(get_db), _u: User = Depends(get_current_user)):
    spu = db.query(Spu).options(joinedload(Spu.skus)).filter(
        Spu.id == spu_id, Spu.deleted_at.is_(None)).first()
    if not spu:
        raise HTTPException(404, "商品不存在或已在回收站")
    return spu_to_out(spu)


def _ensure_code_unique(db: Session, code: str, exclude_id: int | None = None):
    if not code:
        return
    q = db.query(Spu).filter(Spu.code == code, Spu.deleted_at.is_(None))
    if exclude_id:
        q = q.filter(Spu.id != exclude_id)
    if q.first():
        raise HTTPException(400, f"商品编号 {code} 已存在")


@router.post("/spus")
def create_spu(body: SpuCreate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    _ensure_code_unique(db, body.code.strip())
    spu = Spu(
        name=body.name.strip(),
        code=body.code.strip(),
        category_id=_resolve_category(db, body),
        unit=body.unit.strip() or "件",
        weight=body.weight or 0,
        weight_unit=body.weight_unit.strip() or "千克",
        designer=body.designer.strip(),
        production_date=body.production_date,
        material=body.material.strip(),
        image_url=body.image_url.strip(),
        remark=body.remark.strip(),
    )
    for s in body.skus:
        spu.skus.append(Sku(
            spec_name=s.spec_name.strip(),
            sku_code=s.sku_code.strip(),
            barcode=s.barcode.strip(),
            cost_price=s.cost_price or 0,
            sale_price=s.sale_price or 0,
            stock=s.stock or 0,
        ))
    db.add(spu)
    db.flush()
    _sync_images(db, spu, body.images)
    _ensure_unit_name(db, Unit, body.unit)
    _ensure_unit_name(db, WeightUnit, body.weight_unit)
    db.commit()
    db.refresh(spu)
    add_log(db, u, "新增商品", f"{spu.name}（{len(spu.skus)}个规格）")
    db.commit()
    return spu_to_out(spu)


@router.put("/spus/{spu_id}")
def update_spu(spu_id: int, body: SpuUpdate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    spu = db.query(Spu).options(joinedload(Spu.skus)).filter(
        Spu.id == spu_id, Spu.deleted_at.is_(None)).first()
    if not spu:
        raise HTTPException(404, "商品不存在或已在回收站")
    _ensure_code_unique(db, body.code.strip(), exclude_id=spu.id)
    spu.name = body.name.strip()
    spu.code = body.code.strip()
    spu.category_id = _resolve_category(db, body)
    spu.unit = body.unit.strip() or "件"
    spu.weight = body.weight or 0
    spu.weight_unit = body.weight_unit.strip() or "千克"
    spu.designer = body.designer.strip()
    spu.production_date = body.production_date
    spu.material = body.material.strip()
    spu.image_url = body.image_url.strip()
    spu.remark = body.remark.strip()
    # SKU 同步：保留提交中带 id 的、新增无 id 的、删除未提交的
    submitted = [s.id for s in body.skus if s.id]
    for old in list(spu.skus):
        if old.id not in submitted:
            db.delete(old)
    for s in body.skus:
        if s.id:
            target = db.query(Sku).filter(Sku.id == s.id, Sku.spu_id == spu.id).first()
            if target:
                target.spec_name = s.spec_name.strip()
                target.sku_code = s.sku_code.strip()
                target.barcode = s.barcode.strip()
                target.cost_price = s.cost_price or 0
                target.sale_price = s.sale_price or 0
        else:
            spu.skus.append(Sku(
                spec_name=s.spec_name.strip(),
                sku_code=s.sku_code.strip(),
                barcode=s.barcode.strip(),
                cost_price=s.cost_price or 0,
                sale_price=s.sale_price or 0,
                stock=s.stock or 0,
            ))
    _sync_images(db, spu, body.images)
    _ensure_unit_name(db, Unit, body.unit)
    _ensure_unit_name(db, WeightUnit, body.weight_unit)
    db.commit()
    db.refresh(spu)
    add_log(db, u, "修改商品", spu.name)
    db.commit()
    return spu_to_out(spu)


@router.delete("/spus/{spu_id}")
def delete_spu(spu_id: int, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    spu = db.query(Spu).filter(Spu.id == spu_id, Spu.deleted_at.is_(None)).first()
    if not spu:
        raise HTTPException(404, "商品不存在或已在回收站")
    name = spu.name
    spu.deleted_at = datetime.utcnow()  # v1.3：删除进回收站（软删除，历史流水保留）
    db.commit()
    add_log(db, u, "删除商品进回收站", name)
    db.commit()
    return {"ok": True}


# ================= 秒账 Excel / CSV 导入 =================
COLUMN_ALIASES = {
    "code": ["编号", "商品编号", "货号", "编码", "code", "商品编码"],
    "name": ["名称", "品名", "商品名称", "商品", "name"],
    "spec": ["规格", "规格名", "规格名称", "型号", "颜色", "尺寸", "spec", "属性"],
    "barcode": ["条码", "条形码", "一维码", "barcode", "商品条码"],
    "cost": ["进价", "成本价", "成本", "进货价", "cost", "采购价"],
    "price": ["售价", "销售价", "价格", "单价", "卖价", "price", "零售价"],
    "stock": ["库存", "数量", "库存数量", "stock", "quantity", "现存数量"],
    "category": ["分类", "类别", "category", "商品分类"],
    "unit": ["单位", "unit", "计量单位"],
}


def _map_col(header):
    hm = {h: i for i, h in enumerate(header)}
    m = {}
    for key, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            if a in hm:
                m[key] = hm[a]
                break
    return m


def _f(row, idx):
    if idx is None:
        return None
    try:
        v = row[idx]
    except (IndexError, TypeError):
        return None
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _num(row, idx):
    v = _f(row, idx)
    if v is None:
        return 0
    try:
        return float(v)
    except ValueError:
        return 0


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    filename = (file.filename or "").lower()
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "文件为空")
    try:
        if filename.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            rows = [r for r in csv.reader(io.StringIO(text))]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
    except Exception:
        raise HTTPException(400, "文件解析失败，请上传 .xlsx 或 .csv 格式")
    if len(rows) < 2:
        raise HTTPException(400, "文件没有数据行")
    header = [(str(c).strip() if c is not None else "") for c in rows[0]]
    m = _map_col(header)
    if "name" not in m and "code" not in m:
        raise HTTPException(400, "未识别到「名称/编号」列，请检查表头是否包含常见列名")

    # 按 编号 或 名称 分组：同一商品多行 = 多个规格(SKU)
    groups = {}
    for r in rows[1:]:
        code = _f(r, m.get("code"))
        name = _f(r, m.get("name"))
        key = code or name
        if not key:
            continue
        groups.setdefault(key, []).append(r)

    created_spu = 0
    created_sku = 0
    skipped = []
    for key, g in groups.items():
        first = g[0]
        name = _f(first, m.get("name")) or key
        code = _f(first, m.get("code")) or ""
        if code and db.query(Spu).filter(Spu.code == code, Spu.deleted_at.is_(None)).first():
            skipped.append(f"编号[{code}] 已存在，跳过")
            continue
        if not code and db.query(Spu).filter(Spu.name == name, Spu.deleted_at.is_(None)).first():
            skipped.append(f"商品[{name}] 已存在，跳过")
            continue
        spu = Spu(name=name, code=code)
        cat_name = _f(first, m.get("category"))
        if cat_name:
            cat = db.query(Category).filter(Category.name == cat_name).first()
            if not cat:
                cat = Category(name=cat_name)
                db.add(cat)
                db.flush()
            spu.category_id = cat.id
        unit = _f(first, m.get("unit"))
        if unit:
            spu.unit = unit
        single = len(g) == 1
        for r in g:
            spec = _f(r, m.get("spec")) or "默认"
            sku = Sku(
                spec_name=spec,
                # 同商品多规格时 SKU 编码留空串（Pydantic str 不允许 None）
                sku_code=(code or name) if single else "",
                barcode=_f(r, m.get("barcode")) or "",
                cost_price=_num(r, m.get("cost")),
                sale_price=_num(r, m.get("price")),
                stock=int(_num(r, m.get("stock"))),
            )
            spu.skus.append(sku)
            created_sku += 1
        db.add(spu)
        created_spu += 1
    db.commit()
    add_log(db, u, "秒账导入", f"新建商品{created_spu}个 / 规格{created_sku}个 / 跳过{len(skipped)}条")
    db.commit()
    return {
        "created_spu": created_spu,
        "created_sku": created_sku,
        "skipped": skipped[:30],
        "skipped_count": len(skipped),
    }


# ================= 单位字典（计数单位 + 重量单位，可自由增删改） =================

@router.get("/units")
def list_units(db: Session = Depends(get_db), _u: User = Depends(get_current_user)):
    return [UnitOut(id=u.id, name=u.name) for u in db.query(Unit).order_by(Unit.id).all()]


@router.post("/units")
def create_unit(body: UnitCreate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "单位名称不能为空")
    if db.query(Unit).filter(Unit.name == name).first():
        raise HTTPException(400, f"单位「{name}」已存在")
    item = Unit(name=name)
    db.add(item)
    db.commit()
    db.refresh(item)
    add_log(db, u, "新增单位", name)
    db.commit()
    return UnitOut(id=item.id, name=item.name)


@router.put("/units/{uid}")
def update_unit(uid: int, body: UnitUpdate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    item = db.query(Unit).filter(Unit.id == uid).first()
    if not item:
        raise HTTPException(404, "单位不存在")
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "单位名称不能为空")
    dup = db.query(Unit).filter(Unit.name == name, Unit.id != uid).first()
    if dup:
        raise HTTPException(400, f"单位「{name}」已存在")
    old = item.name
    item.name = name
    db.commit()
    add_log(db, u, "修改单位", f"{old} -> {name}")
    db.commit()
    return UnitOut(id=item.id, name=item.name)


@router.delete("/units/{uid}")
def delete_unit(uid: int, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    item = db.query(Unit).filter(Unit.id == uid).first()
    if not item:
        raise HTTPException(404, "单位不存在")
    if db.query(Spu).filter(Spu.unit == item.name).first():
        raise HTTPException(400, f"「{item.name}」正在被商品使用，不能删除")
    name = item.name
    db.delete(item)
    db.commit()
    add_log(db, u, "删除单位", name)
    db.commit()
    return {"ok": True}


@router.get("/weight-units")
def list_weight_units(db: Session = Depends(get_db), _u: User = Depends(get_current_user)):
    return [UnitOut(id=w.id, name=w.name) for w in db.query(WeightUnit).order_by(WeightUnit.id).all()]


@router.post("/weight-units")
def create_weight_unit(body: UnitCreate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "重量单位名称不能为空")
    if db.query(WeightUnit).filter(WeightUnit.name == name).first():
        raise HTTPException(400, f"重量单位「{name}」已存在")
    item = WeightUnit(name=name)
    db.add(item)
    db.commit()
    db.refresh(item)
    add_log(db, u, "新增重量单位", name)
    db.commit()
    return UnitOut(id=item.id, name=item.name)


@router.put("/weight-units/{uid}")
def update_weight_unit(uid: int, body: UnitUpdate, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    item = db.query(WeightUnit).filter(WeightUnit.id == uid).first()
    if not item:
        raise HTTPException(404, "重量单位不存在")
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "重量单位名称不能为空")
    dup = db.query(WeightUnit).filter(WeightUnit.name == name, WeightUnit.id != uid).first()
    if dup:
        raise HTTPException(400, f"重量单位「{name}」已存在")
    old = item.name
    item.name = name
    db.commit()
    add_log(db, u, "修改重量单位", f"{old} -> {name}")
    db.commit()
    return UnitOut(id=item.id, name=item.name)


@router.delete("/weight-units/{uid}")
def delete_weight_unit(uid: int, db: Session = Depends(get_db), u: User = Depends(get_current_user)):
    item = db.query(WeightUnit).filter(WeightUnit.id == uid).first()
    if not item:
        raise HTTPException(404, "重量单位不存在")
    if db.query(Spu).filter(Spu.weight_unit == item.name).first():
        raise HTTPException(400, f"「{item.name}」正在被商品使用，不能删除")
    name = item.name
    db.delete(item)
    db.commit()
    add_log(db, u, "删除重量单位", name)
    db.commit()
    return {"ok": True}
